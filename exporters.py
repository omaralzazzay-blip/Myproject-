# -*- coding: utf-8 -*-
"""تصدير التقارير بصيغتي PDF و Excel مع دعم كامل للعربية (RTL)"""
import os
import io
import arabic_reshaper

# --- استبدال bidi بدالة محلية ---
def get_display(text):
    """إعادة تشكيل النص العربي وعكسه للعرض الصحيح (محاكاة bidi)"""
    if text is None:
        return ''
    return arabic_reshaper.reshape(str(text))[::-1]
# ---------------------------------

from reportlab.lib.pagesizes import A4, landscape as rl_landscape
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from openpyxl import Workbook
from openpyxl.styles import Font as XLFont, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.abspath(__file__))
FONT_DIR = os.path.join(BASE, 'assets', 'fonts')
_loaded = False


def _ensure_fonts():
    global _loaded
    if _loaded:
        return
    pdfmetrics.registerFont(TTFont('NotoAr', os.path.join(FONT_DIR, 'NotoSansArabic-Regular.ttf')))
    pdfmetrics.registerFont(TTFont('NotoArB', os.path.join(FONT_DIR, 'NotoSansArabic-Bold.ttf')))
    _loaded = True


def ar(t):
    """تقويم النص العربي (إعادة تشكيل + اتجاه RTL) لعرضه في PDF."""
    if t is None:
        return ''
    return get_display(str(t))


def make_pdf(title, headers, rows, landscape=False):
    _ensure_fonts()
    buf = io.BytesIO()
    page = rl_landscape(A4) if landscape else A4
    doc = SimpleDocTemplate(buf, pagesize=page,
                            rightMargin=12 * mm, leftMargin=12 * mm,
                            topMargin=14 * mm, bottomMargin=12 * mm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('t', parent=styles['Title'], fontName='NotoArB',
                                 fontSize=15, leading=21, alignment=1)
    cell_style = ParagraphStyle('c', fontName='NotoAr', fontSize=8.5, leading=12, alignment=2)
    head_style = ParagraphStyle('h', parent=cell_style, fontName='NotoArB',
                                textColor=colors.white)
    data = [[Paragraph(ar(h), head_style) for h in headers]]
    for r in rows:
        data.append([Paragraph(ar('' if c is None else str(c)), cell_style) for c in r])
    tbl = Table(data, repeatRows=1)
    tbl.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a5f')),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#cbd5e1')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f2f6fb')]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 5),
        ('RIGHTPADDING', (0, 0), (-1, -1), 5),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    doc.build([Paragraph(ar(title), title_style), Spacer(1, 8), tbl])
    buf.seek(0)
    return buf


def make_xlsx(title, headers, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = 'تقرير'
    ws.sheet_view.rightToLeft = True
    ws.append([title])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.cell(row=1, column=1).font = XLFont(bold=True, size=14, color='1e3a5f')
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=2, column=c)
        cell.font = XLFont(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='1E3A5F')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    for r in rows:
        ws.append(['' if v is None else str(v) for v in r])
    # عرض الأعمدة حسب أطول محتوى (مع حد أقصى)
    for c in range(1, len(headers) + 1):
        width = max([len(str(ws.cell(row=r, column=c).value or '')) for r in range(1, ws.max_row + 1)])
        ws.column_dimensions[get_column_letter(c)].width = min(max(width + 4, 10), 40)
    # محاذاة RTL لكل الخلايا
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, max_col=len(headers)):
        for cell in row:
            cell.alignment = Alignment(horizontal='right', vertical='center')
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf